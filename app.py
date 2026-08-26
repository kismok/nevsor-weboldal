import os
import sqlite3
import threading
import asyncio
import re
import unicodedata
import io

from openpyxl import Workbook

from datetime import datetime

import discord
from discord.ext import commands

from flask import (
    Flask,
    g,
    render_template,
    request,
    redirect,
    url_for,
    jsonify,
    abort,
    session,
    send_file
)


# =========================================================
# FLASK
# =========================================================

app = Flask(__name__)

app.secret_key = os.environ.get(
    "SECRET_KEY",
    "csereld-le-egy-hosszu-titkos-kulcsra"
)


# =========================================================
# BEJELENTKEZÉS
# =========================================================

ADMIN_USERNAME = os.environ.get(
    "ADMIN_USERNAME",
    "admin"
)

ADMIN_PASSWORD = os.environ.get(
    "ADMIN_PASSWORD",
    "admin123"
)


def is_logged_in():

    return session.get("logged_in") is True


def require_login():

    return is_logged_in()


# =========================================================
# HÓNAP KEZELÉS
# =========================================================

def current_month():

    return datetime.now().strftime(
        "%Y-%m"
    )


def normalize_month(value):

    if not value:

        return current_month()

    try:

        datetime.strptime(
            value,
            "%Y-%m"
        )

        return value

    except ValueError:

        return current_month()


def get_selected_month():

    return normalize_month(
        request.args.get(
            "month",
            ""
        )
    )


def month_display(month_value):

    try:

        dt = datetime.strptime(
            month_value,
            "%Y-%m"
        )

        months = [
            "január",
            "február",
            "március",
            "április",
            "május",
            "június",
            "július",
            "augusztus",
            "szeptember",
            "október",
            "november",
            "december"
        ]

        return (
            f"{dt.year}. "
            f"{months[dt.month - 1]}"
        )

    except Exception:

        return month_value


# =========================================================
# ADATBÁZIS
# =========================================================

DB_PATH = os.environ.get(
    "DATABASE_PATH",
    "/tmp/nevsor.db"
)


def get_connection():

    conn = sqlite3.connect(
        DB_PATH,
        timeout=30
    )

    conn.row_factory = sqlite3.Row

    conn.execute(
        "PRAGMA foreign_keys = ON"
    )

    return conn


def db():

    if "db" not in g:

        g.db = get_connection()

    return g.db


@app.teardown_appcontext
def close_db(_error=None):

    conn = g.pop(
        "db",
        None
    )

    if conn:

        conn.close()


# =========================================================
# ADATBÁZIS LÉTREHOZÁS
# =========================================================

def init_db():

    conn = get_connection()

    conn.executescript("""
    CREATE TABLE IF NOT EXISTS members (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        character_id TEXT NOT NULL UNIQUE,
        discord_user_id TEXT UNIQUE,
        discord_username TEXT,
        payment_status INTEGER NOT NULL DEFAULT 0,
        payment_date TEXT,
        created_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS penalty_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        member_id INTEGER NOT NULL,
        points INTEGER NOT NULL,
        reason TEXT NOT NULL,
        created_at TEXT NOT NULL,

        FOREIGN KEY(member_id)
            REFERENCES members(id)
            ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS payments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        member_id INTEGER NOT NULL,
        payment_month TEXT NOT NULL,
        payment_status INTEGER NOT NULL DEFAULT 0,
        payment_date TEXT,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL,

        UNIQUE(member_id, payment_month),

        FOREIGN KEY(member_id)
            REFERENCES members(id)
            ON DELETE CASCADE
    );
    """)

    columns = conn.execute(
        "PRAGMA table_info(members)"
    ).fetchall()

    column_names = [
        column["name"]
        for column in columns
    ]

    if "payment_status" not in column_names:

        conn.execute("""
            ALTER TABLE members
            ADD COLUMN payment_status
            INTEGER NOT NULL DEFAULT 0
        """)

    if "payment_date" not in column_names:

        conn.execute("""
            ALTER TABLE members
            ADD COLUMN payment_date TEXT
        """)

    old_members = conn.execute("""
        SELECT
            id,
            payment_status,
            payment_date,
            created_at

        FROM members

        WHERE
            payment_status != 0
            OR payment_date IS NOT NULL
    """).fetchall()

    for member in old_members:

        if member["payment_date"]:

            try:

                payment_month = datetime.strptime(
                    member["payment_date"],
                    "%Y-%m-%d"
                ).strftime(
                    "%Y-%m"
                )

            except ValueError:

                payment_month = current_month()

        else:

            payment_month = current_month()

        existing = conn.execute("""
            SELECT id

            FROM payments

            WHERE
                member_id = ?
                AND payment_month = ?
        """, (
            member["id"],
            payment_month
        )).fetchone()

        if not existing:

            now = datetime.now().isoformat(
                timespec="seconds"
            )

            conn.execute("""
                INSERT INTO payments
                (
                    member_id,
                    payment_month,
                    payment_status,
                    payment_date,
                    created_at,
                    updated_at
                )

                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                member["id"],
                payment_month,
                int(
                    member["payment_status"]
                    or 0
                ),
                member["payment_date"],
                member["created_at"] or now,
                now
            ))

    conn.commit()

    conn.close()


# =========================================================
# BEJELENTKEZÉS
# =========================================================

@app.route(
    "/login",
    methods=["GET", "POST"]
)
def login():

    if request.method == "POST":

        username = request.form.get(
            "username",
            ""
        ).strip()

        password = request.form.get(
            "password",
            ""
        )

        if (
            username == ADMIN_USERNAME
            and password == ADMIN_PASSWORD
        ):

            session["logged_in"] = True

            return redirect(
                url_for("index")
            )

        return render_template(
            "login.html",
            error="Hibás felhasználónév vagy jelszó."
        )

    return render_template(
        "login.html",
        error=None
    )


# =========================================================
# KIJELENTKEZÉS
# =========================================================

@app.route(
    "/logout"
)
def logout():

    session.clear()

    return redirect(
        url_for("index")
    )


# =========================================================
# FŐOLDAL
# =========================================================

@app.route("/")
def index():

    q = request.args.get(
        "q",
        ""
    ).strip()

    selected_month = get_selected_month()

    query = """
        SELECT

            m.id,
            m.name,
            m.character_id,
            m.discord_user_id,
            m.discord_username,
            m.created_at,

            COALESCE(
                pmt.payment_status,
                0
            ) AS payment_status,

            pmt.payment_date,

            COALESCE(
                SUM(p.points),
                0
            ) AS total_points

        FROM members m

        LEFT JOIN payments pmt

            ON pmt.member_id = m.id

            AND pmt.payment_month = ?

        LEFT JOIN penalty_log p

            ON p.member_id = m.id
    """

    params = [
        selected_month
    ]

    if q:

        like = f"%{q}%"

        query += """
            WHERE
                m.name LIKE ?
                OR m.character_id LIKE ?
                OR m.discord_username LIKE ?
        """

        params.extend([
            like,
            like,
            like
        ])

    query += """
        GROUP BY
            m.id,
            m.name,
            m.character_id,
            m.discord_user_id,
            m.discord_username,
            m.created_at,
            pmt.payment_status,
            pmt.payment_date

        ORDER BY
            m.name COLLATE NOCASE
    """

    members = db().execute(
        query,
        params
    ).fetchall()

    return render_template(
        "index.html",
        members=members,
        q=q,
        selected_month=selected_month,
        selected_month_display=month_display(
            selected_month
        ),
        logged_in=is_logged_in()
    )


# =========================================================
# ÚJ TAG
# =========================================================

@app.route(
    "/member/add",
    methods=["POST"]
)
def add_member():

    if not require_login():

        return redirect(
            url_for("login")
        )

    name = request.form.get(
        "name",
        ""
    ).strip()

    character_id = request.form.get(
        "character_id",
        ""
    ).strip()

    payment_status = request.form.get(
        "payment_status",
        "0"
    )

    payment_date = request.form.get(
        "payment_date",
        ""
    ).strip()

    payment_month = normalize_month(
        request.form.get(
            "month",
            ""
        )
    )

    if not name or not character_id:

        return redirect(
            url_for(
                "index",
                month=payment_month
            )
        )

    try:

        status = int(
            payment_status
        )

    except (
        ValueError,
        TypeError
    ):

        status = 0

    if status not in (
        0,
        1,
        2
    ):

        status = 0

    if status == 2:

        payment_date = ""

    now = datetime.now().isoformat(
        timespec="seconds"
    )

    try:

        conn = db()

        cursor = conn.execute("""
            INSERT INTO members
            (
                name,
                character_id,
                created_at
            )

            VALUES (?, ?, ?)
        """, (
            name,
            character_id,
            now
        ))

        member_id = cursor.lastrowid

        conn.execute("""
            INSERT INTO payments
            (
                member_id,
                payment_month,
                payment_status,
                payment_date,
                created_at,
                updated_at
            )

            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            member_id,
            payment_month,
            status,
            payment_date or None,
            now,
            now
        ))

        conn.commit()

    except sqlite3.IntegrityError:

        db().rollback()

    return redirect(
        url_for(
            "index",
            month=payment_month
        )
    )


# =========================================================
# TAG SZERKESZTÉSE
# =========================================================

@app.route(
    "/member/<int:member_id>/edit",
    methods=["POST"]
)
def edit_member(member_id):

    if not require_login():

        return redirect(
            url_for("login")
        )

    name = request.form.get(
        "name",
        ""
    ).strip()

    character_id = request.form.get(
        "character_id",
        ""
    ).strip()

    payment_status = request.form.get(
        "payment_status",
        "0"
    )

    payment_date = request.form.get(
        "payment_date",
        ""
    ).strip()

    payment_month = normalize_month(
        request.form.get(
            "month",
            ""
        )
    )

    if not name or not character_id:

        return redirect(
            url_for(
                "index",
                month=payment_month
            )
        )

    try:

        status = int(
            payment_status
        )

    except (
        ValueError,
        TypeError
    ):

        status = 0

    if status not in (
        0,
        1,
        2
    ):

        status = 0

    if status == 2:

        payment_date = ""

    now = datetime.now().isoformat(
        timespec="seconds"
    )

    try:

        conn = db()

        member_exists = conn.execute("""
            SELECT id
            FROM members
            WHERE id = ?
        """, (
            member_id,
        )).fetchone()

        if not member_exists:

            abort(404)

        conn.execute("""
            UPDATE members

            SET
                name = ?,
                character_id = ?

            WHERE id = ?
        """, (
            name,
            character_id,
            member_id
        ))

        existing = conn.execute("""
            SELECT id

            FROM payments

            WHERE
                member_id = ?
                AND payment_month = ?
        """, (
            member_id,
            payment_month
        )).fetchone()

        if existing:

            conn.execute("""
                UPDATE payments

                SET
                    payment_status = ?,
                    payment_date = ?,
                    updated_at = ?

                WHERE
                    member_id = ?
                    AND payment_month = ?
            """, (
                status,
                payment_date or None,
                now,
                member_id,
                payment_month
            ))

        else:

            conn.execute("""
                INSERT INTO payments
                (
                    member_id,
                    payment_month,
                    payment_status,
                    payment_date,
                    created_at,
                    updated_at
                )

                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                member_id,
                payment_month,
                status,
                payment_date or None,
                now,
                now
            ))

        conn.commit()

    except sqlite3.IntegrityError:

        db().rollback()

    return redirect(
        url_for(
            "index",
            month=payment_month
        )
    )


# =========================================================
# HIBAPONT NAPLÓ OLDAL
# =========================================================

@app.route(
    "/penalty-log"
)
def penalty_log():

    q = request.args.get(
        "q",
        ""
    ).strip()

    query = """
        SELECT
            p.id,
            p.member_id,
            p.points,
            p.reason,
            p.created_at,
            m.name,
            m.character_id

        FROM penalty_log p

        JOIN members m
            ON m.id = p.member_id
    """

    params = []

    if q:

        like = f"%{q}%"

        query += """
            WHERE
                m.name LIKE ?
                OR m.character_id LIKE ?
                OR p.reason LIKE ?
        """

        params.extend([
            like,
            like,
            like
        ])

    query += """
        ORDER BY
            p.created_at DESC,
            p.id DESC
    """

    penalties = db().execute(
        query,
        params
    ).fetchall()

    return render_template(
        "penalty_log.html",
        penalties=penalties,
        q=q,
        logged_in=is_logged_in()
    )


# =========================================================
# TAG ADATAI
# =========================================================

@app.route(
    "/member/<int:member_id>"
)
def member_detail(member_id):

    selected_month = normalize_month(
        request.args.get(
            "month",
            ""
        )
    )

    conn = db()

    member = conn.execute("""
        SELECT

            m.id,
            m.name,
            m.character_id,
            m.discord_user_id,
            m.discord_username,

            COALESCE(
                pmt.payment_status,
                0
            ) AS payment_status,

            pmt.payment_date,

            COALESCE(
                SUM(p.points),
                0
            ) AS total_points

        FROM members m

        LEFT JOIN payments pmt

            ON pmt.member_id = m.id

            AND pmt.payment_month = ?

        LEFT JOIN penalty_log p

            ON p.member_id = m.id

        WHERE m.id = ?

        GROUP BY
            m.id,
            m.name,
            m.character_id,
            m.discord_user_id,
            m.discord_username,
            pmt.payment_status,
            pmt.payment_date
    """, (
        selected_month,
        member_id
    )).fetchone()

    if not member:

        abort(404)

    logs = conn.execute("""
        SELECT *

        FROM penalty_log

        WHERE member_id = ?

        ORDER BY id DESC
    """, (
        member_id,
    )).fetchall()

    return jsonify({

        "id":
            member["id"],

        "name":
            member["name"],

        "character_id":
            member["character_id"],

        "discord_user_id":
            member["discord_user_id"],

        "discord_username":
            member["discord_username"],

        "payment_status":
            int(
                member["payment_status"]
                or 0
            ),

        "payment_date":
            member["payment_date"],

        "payment_month":
            selected_month,

        "total_points":
            int(
                member["total_points"]
                or 0
            ),

        "logged_in":
            is_logged_in(),

        "logs":
            [
                dict(x)
                for x in logs
            ]

    })


# =========================================================
# HIBAPONT HOZZÁADÁSA
# =========================================================

@app.route(
    "/member/<int:member_id>/penalty",
    methods=["POST"]
)
def add_penalty(member_id):

    if not require_login():

        return redirect(
            url_for("login")
        )

    points = request.form.get(
        "points",
        "0"
    ).strip()

    reason = request.form.get(
        "reason",
        ""
    ).strip()

    month = normalize_month(
        request.form.get(
            "month",
            ""
        )
    )

    try:

        points = int(
            points
        )

    except ValueError:

        points = 0

    if points > 0 and reason:

        db().execute("""
            INSERT INTO penalty_log
            (
                member_id,
                points,
                reason,
                created_at
            )

            VALUES (?, ?, ?, ?)
        """, (
            member_id,
            points,
            reason,
            datetime.now().isoformat(
                timespec="seconds"
            )
        ))

        db().commit()

    return redirect(
        url_for(
            "index",
            month=month
        )
    )


# =========================================================
# HIBAPONT TÖRLÉSE
# =========================================================

@app.route(
    "/penalty/<int:penalty_id>/delete",
    methods=["POST"]
)
def delete_penalty(penalty_id):

    if not require_login():

        return redirect(
            url_for("login")
        )

    month = normalize_month(
        request.form.get(
            "month",
            ""
        )
    )

    conn = db()

    penalty = conn.execute("""
        SELECT member_id

        FROM penalty_log

        WHERE id = ?
    """, (
        penalty_id,
    )).fetchone()

    if not penalty:

        abort(404)

    conn.execute("""
        DELETE FROM penalty_log

        WHERE id = ?
    """, (
        penalty_id,
    ))

    conn.commit()

    return redirect(
        url_for(
            "index",
            month=month
        )
    )


# =========================================================
# DISCORD KÉZI BEÁLLÍTÁS
# =========================================================

@app.route(
    "/member/<int:member_id>/discord",
    methods=["POST"]
)
def set_discord(member_id):

    if not require_login():

        return redirect(
            url_for("login")
        )

    discord_user_id = request.form.get(
        "discord_user_id",
        ""
    ).strip()

    discord_username = request.form.get(
        "discord_username",
        ""
    ).strip()

    month = normalize_month(
        request.form.get(
            "month",
            ""
        )
    )

    try:

        db().execute("""
            UPDATE members

            SET
                discord_user_id = ?,
                discord_username = ?

            WHERE id = ?
        """, (
            discord_user_id or None,
            discord_username or None,
            member_id
        ))

        db().commit()

    except sqlite3.IntegrityError:

        db().rollback()

    return redirect(
        url_for(
            "index",
            month=month
        )
    )


# =========================================================
# TAG TÖRLÉSE
# =========================================================

@app.route(
    "/member/<int:member_id>/delete",
    methods=["POST"]
)
def delete_member(member_id):

    if not require_login():

        return redirect(
            url_for("login")
        )

    month = normalize_month(
        request.form.get(
            "month",
            ""
        )
    )

    db().execute("""
        DELETE FROM members

        WHERE id = ?
    """, (
        member_id,
    ))

    db().commit()

    return redirect(
        url_for(
            "index",
            month=month
        )
    )




# =========================================================
# EXCEL EXPORTÁLÁS
# =========================================================

@app.route(
    "/export/excel"
)
def export_database():

    if not require_login():

        return redirect(
            url_for("login")
        )

    conn = get_connection()

    try:

        members = conn.execute("""
            SELECT
                id,
                name,
                character_id,
                discord_user_id,
                discord_username,
                created_at

            FROM members

            ORDER BY name COLLATE NOCASE
        """).fetchall()

        payments = conn.execute("""
            SELECT
                p.id,
                p.member_id,
                m.name,
                m.character_id,
                p.payment_month,
                p.payment_status,
                p.payment_date,
                p.created_at,
                p.updated_at

            FROM payments p

            JOIN members m
                ON m.id = p.member_id

            ORDER BY
                p.payment_month DESC,
                m.name COLLATE NOCASE
        """).fetchall()

        penalties = conn.execute("""
            SELECT
                p.id,
                p.member_id,
                m.name,
                m.character_id,
                p.points,
                p.reason,
                p.created_at

            FROM penalty_log p

            JOIN members m
                ON m.id = p.member_id

            ORDER BY
                p.created_at DESC,
                p.id DESC
        """).fetchall()

    finally:

        conn.close()

    workbook = Workbook()

    members_sheet = workbook.active
    members_sheet.title = "Tagok"

    members_sheet.append([
        "ID",
        "Név",
        "Karakter ID",
        "Discord felhasználó ID",
        "Discord név",
        "Létrehozva"
    ])

    for member in members:

        members_sheet.append([
            member["id"],
            member["name"],
            member["character_id"],
            member["discord_user_id"],
            member["discord_username"],
            member["created_at"]
        ])

    payments_sheet = workbook.create_sheet(
        "Befizetések"
    )

    payments_sheet.append([
        "Befizetés ID",
        "Tag ID",
        "Név",
        "Karakter ID",
        "Hónap",
        "Státusz",
        "Befizetés dátuma",
        "Létrehozva",
        "Módosítva"
    ])

    status_names = {
        0: "Nincs befizetve",
        1: "Befizetve",
        2: "Felmentve"
    }

    for payment in payments:

        payments_sheet.append([
            payment["id"],
            payment["member_id"],
            payment["name"],
            payment["character_id"],
            payment["payment_month"],
            status_names.get(
                int(payment["payment_status"] or 0),
                "Ismeretlen"
            ),
            payment["payment_date"],
            payment["created_at"],
            payment["updated_at"]
        ])

    penalties_sheet = workbook.create_sheet(
        "Hibapontok"
    )

    penalties_sheet.append([
        "Hibapont ID",
        "Tag ID",
        "Név",
        "Karakter ID",
        "Pont",
        "Indok",
        "Dátum"
    ])

    for penalty in penalties:

        penalties_sheet.append([
            penalty["id"],
            penalty["member_id"],
            penalty["name"],
            penalty["character_id"],
            penalty["points"],
            penalty["reason"],
            penalty["created_at"]
        ])

    for sheet in workbook.worksheets:

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        for column_cells in sheet.columns:

            max_length = 0

            for cell in column_cells:

                value = "" if cell.value is None else str(cell.value)

                if len(value) > max_length:

                    max_length = len(value)

            column_letter = column_cells[0].column_letter

            sheet.column_dimensions[column_letter].width = min(
                max(max_length + 2, 10),
                50
            )

    buffer = io.BytesIO()

    workbook.save(
        buffer
    )

    buffer.seek(0)

    filename = (
        f"nevsor-export-"
        f"{datetime.now().strftime('%Y-%m-%d-%H%M%S')}.xlsx"
    )

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )


# =========================================================
# DISCORD BOT
# =========================================================

DISCORD_TOKEN = os.environ.get(
    "DISCORD_TOKEN"
)

DISCORD_GUILD_ID = os.environ.get(
    "DISCORD_GUILD_ID"
)


discord_loop = None
discord_started = False


intents = discord.Intents.default()

# Discord Developer Portalban is engedélyezni kell:
# SERVER MEMBERS INTENT
# MESSAGE CONTENT INTENT

intents.members = True
intents.message_content = True


bot = commands.Bot(
    command_prefix="/",
    intents=intents
)


@bot.event
async def on_ready():

    global discord_loop

    discord_loop = asyncio.get_running_loop()

    print(
        f"✅ Discord bot elindult: {bot.user}"
    )

    print(
        f"📡 Bot ID: {bot.user.id}"
    )

    if DISCORD_GUILD_ID:

        print(
            f"🏠 Discord szerver ID: "
            f"{DISCORD_GUILD_ID}"
        )

    else:

        print(
            "❌ DISCORD_GUILD_ID nincs beállítva!"
        )


# =========================================================
# /LINK KARAKTER_ID
# =========================================================

@bot.command(
    name="link"
)
async def link_character(
    ctx,
    character_id=None
):

    if not character_id:

        await ctx.send(
            "❌ Használat: /link KARAKTER_ID"
        )

        return

    conn = get_connection()

    try:

        member = conn.execute("""
            SELECT
                id,
                name,
                character_id

            FROM members

            WHERE character_id = ?
        """, (
            str(character_id),
        )).fetchone()

        if not member:

            await ctx.send(
                f"❌ Nem található karakter "
                f"ezzel az ID-val: `{character_id}`"
            )

            return

        existing = conn.execute("""
            SELECT
                name,
                character_id

            FROM members

            WHERE discord_user_id = ?
        """, (
            str(ctx.author.id),
        )).fetchone()

        if existing:

            await ctx.send(
                "⚠️ Ez a Discord fiók már össze van kapcsolva "
                f"ezzel a karakterrel: "
                f"**{existing['name']}** "
                f"(ID: `{existing['character_id']}`)"
            )

            return

        conn.execute("""
            UPDATE members

            SET
                discord_user_id = ?,
                discord_username = ?

            WHERE id = ?
        """, (
            str(ctx.author.id),
            ctx.author.display_name,
            member["id"]
        ))

        conn.commit()

        await ctx.send(
            "✅ **Sikeres összekapcsolás!**\n\n"
            f"👤 Karakter: **{member['name']}**\n"
            f"🆔 Karakter ID: `{member['character_id']}`\n"
            f"💬 Discord: **{ctx.author.display_name}**"
        )

    except sqlite3.IntegrityError:

        conn.rollback()

        await ctx.send(
            "❌ Ez a Discord fiók vagy karakter "
            "már hozzá van rendelve valakihez."
        )

    finally:

        conn.close()


# =========================================================
# DISCORD NÉV NORMALIZÁLÁS
# =========================================================

def normalize_name(value):

    if not value:

        return ""

    value = str(
        value
    )

    value = value.strip()
    value = value.lstrip("@")
    value = value.casefold()

    value = unicodedata.normalize(
        "NFKD",
        value
    )

    value = "".join(
        char
        for char in value
        if not unicodedata.combining(
            char
        )
    )

    value = re.sub(
        r"[^a-z0-9]+",
        "",
        value
    )

    return value


# =========================================================
# DISCORD NÉV EGYEZÉS
# =========================================================

def names_match(
    db_name,
    discord_member
):

    normalized_db_name = normalize_name(
        db_name
    )

    if not normalized_db_name:

        return False

    possible_names = [

        getattr(
            discord_member,
            "name",
            None
        ),

        getattr(
            discord_member,
            "display_name",
            None
        ),

        getattr(
            discord_member,
            "global_name",
            None
        ),

        str(
            discord_member
        )

    ]

    for discord_name in possible_names:

        normalized_discord_name = normalize_name(
            discord_name
        )

        if not normalized_discord_name:

            continue

        if (
            normalized_discord_name
            == normalized_db_name
        ):

            return True

    return False


# =========================================================
# TELJES DISCORD TAGLISTA LEKÉRÉSE
# =========================================================

async def get_discord_members(guild):

    print(
        "📡 Teljes Discord taglista lekérése..."
    )

    fetched_members = []

    try:

        async for member in guild.fetch_members(
            limit=None
        ):

            if member.bot:

                continue

            fetched_members.append(
                member
            )

        print(
            f"✅ Lekért Discord tagok száma: "
            f"{len(fetched_members)}"
        )

        return fetched_members

    except Exception as e:

        print(
            f"❌ Discord taglista lekérési hiba: "
            f"{repr(e)}"
        )

        print(
            "⚠️ Cache-ben lévő Discord tagok használata..."
        )

        cached_members = [

            member

            for member in guild.members

            if not member.bot

        ]

        print(
            f"⚠️ Cache-ben lévő tagok száma: "
            f"{len(cached_members)}"
        )

        return cached_members


# =========================================================
# DISCORD SZERVER SZINKRONIZÁLÁS
# =========================================================

async def sync_discord_members():

    if not DISCORD_GUILD_ID:

        return {

            "success": False,

            "error":
                "DISCORD_GUILD_ID nincs beállítva."

        }

    try:

        guild_id = int(
            DISCORD_GUILD_ID
        )

    except (
        ValueError,
        TypeError
    ):

        return {

            "success": False,

            "error":
                "Hibás DISCORD_GUILD_ID."

        }

    guild = bot.get_guild(
        guild_id
    )

    if not guild:

        return {

            "success": False,

            "error":
                "A Discord szerver nem található."

        }

    discord_members = await get_discord_members(
        guild
    )

    if not discord_members:

        return {

            "success": False,

            "error":
                "Nem sikerült lekérni a Discord szerver taglistáját."

        }

    conn = get_connection()

    try:

        members_db = conn.execute("""
            SELECT
                id,
                name,
                discord_user_id,
                discord_username

            FROM members
        """).fetchall()

        linked = []

        not_found = []

        used_discord_ids = set()

        for db_member in members_db:

            found_member = None

            # =================================================
            # 1. KORÁBBAN ELMENTETT DISCORD ID
            # =================================================

            if db_member["discord_user_id"]:

                try:

                    discord_id = int(
                        db_member[
                            "discord_user_id"
                        ]
                    )

                    for discord_member in discord_members:

                        if (
                            discord_member.id
                            == discord_id
                        ):

                            found_member = discord_member

                            break

                except (
                    ValueError,
                    TypeError
                ):

                    pass

            # =================================================
            # 2. ELMENTETT DISCORD NÉV
            # =================================================

            if (
                not found_member
                and db_member["discord_username"]
            ):

                saved_name = normalize_name(
                    db_member[
                        "discord_username"
                    ]
                )

                for discord_member in discord_members:

                    if (
                        discord_member.id
                        in used_discord_ids
                    ):

                        continue

                    possible_names = [

                        discord_member.name,

                        discord_member.display_name,

                        getattr(
                            discord_member,
                            "global_name",
                            None
                        )

                    ]

                    for possible_name in possible_names:

                        if not possible_name:

                            continue

                        if (
                            normalize_name(
                                possible_name
                            )
                            == saved_name
                        ):

                            found_member = discord_member

                            break

                    if found_member:

                        break

            # =================================================
            # 3. NÉVSORBAN LÉVŐ NÉV ALAPJÁN
            # =================================================

            if not found_member:

                for discord_member in discord_members:

                    if (
                        discord_member.id
                        in used_discord_ids
                    ):

                        continue

                    if names_match(
                        db_member["name"],
                        discord_member
                    ):

                        found_member = discord_member

                        break

            # =================================================
            # TALÁLAT
            # =================================================

            if found_member:

                used_discord_ids.add(
                    found_member.id
                )

                conn.execute("""
                    UPDATE members

                    SET
                        discord_user_id = ?,
                        discord_username = ?

                    WHERE id = ?
                """, (
                    str(
                        found_member.id
                    ),
                    found_member.display_name,
                    db_member["id"]
                ))

                linked.append({

                    "member_id":
                        db_member["id"],

                    "name":
                        db_member["name"],

                    "discord":
                        found_member.display_name

                })

            else:

                not_found.append({

                    "member_id":
                        db_member["id"],

                    "name":
                        db_member["name"]

                })

        conn.commit()

        discord_only = []

        for discord_member in discord_members:

            if (
                discord_member.id
                not in used_discord_ids
            ):

                discord_only.append({

                    "id":
                        str(
                            discord_member.id
                        ),

                    "username":
                        discord_member.name,

                    "display_name":
                        discord_member.display_name

                })

        return {

            "success": True,

            "total":
                len(
                    discord_members
                ),

            "matched":
                len(
                    linked
                ),

            "unmatched":
                len(
                    discord_only
                ),

            "linked":
                linked,

            "not_found":
                not_found,

            "discord_only":
                discord_only

        }

    except Exception as e:

        conn.rollback()

        print(
            f"❌ Discord szinkron hiba: "
            f"{repr(e)}"
        )

        return {

            "success": False,

            "error":
                f"Discord szinkron hiba: "
                f"{repr(e)}"

        }

    finally:

        conn.close()


# =========================================================
# API - DISCORD SZINKRON
# =========================================================

@app.route(
    "/discord/sync",
    methods=["POST"]
)
def discord_sync():

    global discord_loop

    if not require_login():

        return jsonify({

            "success": False,

            "error":
                "Bejelentkezés szükséges."

        }), 401

    if not DISCORD_TOKEN:

        return jsonify({

            "success": False,

            "error":
                "DISCORD_TOKEN nincs beállítva."

        }), 503

    if not DISCORD_GUILD_ID:

        return jsonify({

            "success": False,

            "error":
                "DISCORD_GUILD_ID nincs beállítva."

        }), 503

    if not discord_started:

        return jsonify({

            "success": False,

            "error":
                "A Discord bot még nem indult el."

        }), 503

    if not bot.is_ready():

        return jsonify({

            "success": False,

            "error":
                "A Discord bot még nincs csatlakozva."

        }), 503

    if discord_loop is None:

        return jsonify({

            "success": False,

            "error":
                "A Discord bot eseményhurka még nem indult el."

        }), 503

    try:

        future = asyncio.run_coroutine_threadsafe(
            sync_discord_members(),
            discord_loop
        )

        result = future.result(
            timeout=90
        )

        if not result.get(
            "success"
        ):

            return jsonify(
                result
            ), 500

        return jsonify(
            result
        )

    except TimeoutError:

        print(
            "❌ Discord sync időtúllépés."
        )

        return jsonify({

            "success": False,

            "error":
                "A Discord szerver válasza túl sokáig tartott."

        }), 504

    except Exception as e:

        print(
            f"❌ Discord sync hiba: "
            f"{repr(e)}"
        )

        return jsonify({

            "success": False,

            "error":
                f"Discord sync hiba: "
                f"{repr(e)}"

        }), 500


# =========================================================
# API - DISCORDON VAN, DE NINCS A NÉVSORBAN
# =========================================================

@app.route(
    "/discord/unmatched"
)
def discord_unmatched():

    if not require_login():

        return jsonify({

            "members": []

        })

    if not bot.is_ready():

        return jsonify({

            "members": []

        })

    if not DISCORD_GUILD_ID:

        return jsonify({

            "members": []

        })

    if discord_loop is None:

        return jsonify({

            "members": []

        })

    async def get_unmatched():

        try:

            guild_id = int(
                DISCORD_GUILD_ID
            )

        except (
            ValueError,
            TypeError
        ):

            return []

        guild = bot.get_guild(
            guild_id
        )

        if not guild:

            return []

        # TELJES TAGLISTA LEKÉRÉSE

        discord_members = await get_discord_members(
            guild
        )

        conn = get_connection()

        try:

            db_members = conn.execute("""
                SELECT
                    name,
                    discord_user_id,
                    discord_username

                FROM members
            """).fetchall()

        finally:

            conn.close()

        used_ids = set()

        for db_member in db_members:

            found_member = None

            # 1. Discord ID

            if db_member["discord_user_id"]:

                try:

                    discord_id = int(
                        db_member[
                            "discord_user_id"
                        ]
                    )

                    for discord_member in discord_members:

                        if (
                            discord_member.id
                            == discord_id
                        ):

                            found_member = discord_member

                            break

                except (
                    ValueError,
                    TypeError
                ):

                    pass

            # 2. Mentett Discord név

            if (
                not found_member
                and db_member["discord_username"]
            ):

                saved_name = normalize_name(
                    db_member[
                        "discord_username"
                    ]
                )

                for discord_member in discord_members:

                    possible_names = [

                        discord_member.name,

                        discord_member.display_name,

                        getattr(
                            discord_member,
                            "global_name",
                            None
                        )

                    ]

                    if any(

                        normalize_name(
                            x
                        )
                        == saved_name

                        for x in possible_names

                        if x

                    ):

                        found_member = discord_member

                        break

            # 3. Név alapján

            if not found_member:

                for discord_member in discord_members:

                    if names_match(
                        db_member["name"],
                        discord_member
                    ):

                        found_member = discord_member

                        break

            if found_member:

                used_ids.add(
                    found_member.id
                )

        result = []

        for discord_member in discord_members:

            if (
                discord_member.id
                not in used_ids
            ):

                result.append({

                    "id":
                        str(
                            discord_member.id
                        ),

                    "username":
                        discord_member.name,

                    "display_name":
                        discord_member.display_name

                })

        return result

    try:

        future = asyncio.run_coroutine_threadsafe(
            get_unmatched(),
            discord_loop
        )

        members = future.result(
            timeout=90
        )

        return jsonify({

            "members":
                members

        })

    except TimeoutError:

        print(
            "❌ Discord unmatched időtúllépés."
        )

        return jsonify({

            "members": []

        })

    except Exception as e:

        print(
            f"❌ Discord unmatched hiba: "
            f"{repr(e)}"
        )

        return jsonify({

            "members": []

        })


# =========================================================
# DISCORD BOT INDÍTÁSA
# =========================================================

def run_discord_bot():

    global discord_loop

    if not DISCORD_TOKEN:

        print(
            "❌ DISCORD_TOKEN nincs beállítva!"
        )

        return

    async def start_bot():

        global discord_loop

        discord_loop = asyncio.get_running_loop()

        print(
            "🚀 Discord bot csatlakoztatása..."
        )

        await bot.start(
            DISCORD_TOKEN
        )

    try:

        asyncio.run(
            start_bot()
        )

    except Exception as e:

        print(
            f"❌ Discord bot hiba: "
            f"{repr(e)}"
        )


def start_discord_bot():

    global discord_started

    if discord_started:

        return

    if not DISCORD_TOKEN:

        print(
            "❌ DISCORD_TOKEN nincs beállítva!"
        )

        return

    discord_started = True

    print(
        "🚀 Discord bot indítása..."
    )

    discord_thread = threading.Thread(
        target=run_discord_bot,
        daemon=True
    )

    discord_thread.start()


# =========================================================
# ALKALMAZÁS INDÍTÁS ELŐTT
# =========================================================

init_db()

start_discord_bot()


# =========================================================
# PROGRAM INDÍTÁSA
# =========================================================

if __name__ == "__main__":

    port = int(
        os.environ.get(
            "PORT",
            "3000"
        )
    )

    app.run(
        host="0.0.0.0",
        port=port,
        debug=False
    )
